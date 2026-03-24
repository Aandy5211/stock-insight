"""Excel 报告导出（openpyxl）"""
import io
import logging
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

logger = logging.getLogger(__name__)

# ── 样式常量 ──────────────────────────────────────────────────────────────────
_RED   = "FFD62728"
_GREEN = "FF2CA02C"
_BLUE  = "FF1F77B4"
_GOLD  = "FFFFD700"
_LGRAY = "FFF5F5F5"
_DGRAY = "FF666666"
_WHITE = "FFFFFFFF"

_HEADER_FILL = PatternFill("solid", fgColor="FF1F4E79")
_SUBHDR_FILL = PatternFill("solid", fgColor="FF2E75B6")
_ALT_FILL    = PatternFill("solid", fgColor=_LGRAY)

_HEADER_FONT = Font(name="微软雅黑", bold=True, color=_WHITE, size=10)
_TITLE_FONT  = Font(name="微软雅黑", bold=True, color="FF1F4E79", size=14)
_BOLD_FONT   = Font(name="微软雅黑", bold=True, size=9)
_NORMAL_FONT = Font(name="微软雅黑", size=9)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_RIGHT  = Alignment(horizontal="right",  vertical="center")

_THIN_BORDER = Border(
    left=Side(style="thin", color=_DGRAY),
    right=Side(style="thin", color=_DGRAY),
    top=Side(style="thin", color=_DGRAY),
    bottom=Side(style="thin", color=_DGRAY),
)


# ── 工具函数 ──────────────────────────────────────────────────────────────────

def _auto_width(ws, min_w: int = 8, max_w: int = 40):
    """自动调整列宽"""
    for col_cells in ws.columns:
        length = max(
            (len(str(cell.value or "")) for cell in col_cells),
            default=min_w,
        )
        # 中文字符宽度约为2倍
        length = min(max(length * 1.4, min_w), max_w)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = length


def _write_header_row(ws, row: int, headers: list[str], fill=None):
    """写表头行"""
    fill = fill or _HEADER_FILL
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    ws.row_dimensions[row].height = 20


def _write_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1,
                        title: str = "", highlight_neg_cols: list = None):
    """将 DataFrame 写入工作表，支持标题行和负数变色"""
    r = start_row
    if title:
        ws.cell(row=r, column=1, value=title).font = _BOLD_FONT
        r += 1

    _write_header_row(ws, r, list(df.columns))
    r += 1

    for i, (_, row_data) in enumerate(df.iterrows()):
        fill = _ALT_FILL if i % 2 == 0 else None
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _NORMAL_FONT
            cell.alignment = _CENTER if col > 1 else _LEFT
            cell.border = _THIN_BORDER
            if fill:
                cell.fill = fill
            # 负数绿色 / 正数红色
            if highlight_neg_cols:
                col_name = df.columns[col - 1]
                if col_name in highlight_neg_cols:
                    try:
                        fv = float(val)
                        cell.font = Font(
                            name="微软雅黑", size=9, bold=True,
                            color=_RED if fv >= 0 else _GREEN,
                        )
                    except Exception:
                        pass
        r += 1
    return r


def _write_kv_block(ws, items: list[tuple], start_row: int, start_col: int = 1):
    """写 key-value 块"""
    r = start_row
    for key, val in items:
        k_cell = ws.cell(row=r, column=start_col, value=str(key))
        k_cell.font = _BOLD_FONT
        k_cell.fill = _ALT_FILL
        k_cell.alignment = _LEFT
        k_cell.border = _THIN_BORDER
        v_cell = ws.cell(row=r, column=start_col + 1, value=val)
        v_cell.font = _NORMAL_FONT
        v_cell.alignment = _LEFT
        v_cell.border = _THIN_BORDER
        r += 1
    return r


# ── 各工作表构建 ──────────────────────────────────────────────────────────────

def _build_summary_sheet(wb, stock_name: str, code: str,
                          quote_items: list, score_items: list,
                          valuation_items: list):
    ws = wb.active
    ws.title = "摘要"
    ws.sheet_view.showGridLines = False

    # 主标题
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"StockInsight 研究报告 — {stock_name}（{code}）"
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _CENTER
    ws.row_dimensions[1].height = 30

    # 生成时间
    ws.merge_cells("A2:F2")
    ws["A2"].value = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="微软雅黑", size=9, color=_DGRAY)
    ws["A2"].alignment = _CENTER

    # 行情信息
    ws.cell(row=4, column=1).value = "行情概览"
    ws.cell(row=4, column=1).font = _BOLD_FONT
    ws.cell(row=4, column=1).fill = _SUBHDR_FILL
    ws.cell(row=4, column=1).font = _HEADER_FONT
    _write_kv_block(ws, quote_items, start_row=5, start_col=1)

    # 财务评分
    ws.cell(row=4, column=3).value = "财务评分"
    ws.cell(row=4, column=3).font = _HEADER_FONT
    ws.cell(row=4, column=3).fill = _SUBHDR_FILL
    _write_kv_block(ws, score_items, start_row=5, start_col=3)

    # 估值评分
    ws.cell(row=4, column=5).value = "估值评分"
    ws.cell(row=4, column=5).font = _HEADER_FONT
    ws.cell(row=4, column=5).fill = _SUBHDR_FILL
    _write_kv_block(ws, valuation_items, start_row=5, start_col=5)

    _auto_width(ws)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18


def _build_financial_sheet(wb, df: pd.DataFrame, sheet_name: str,
                             highlight_cols: list = None):
    if df is None or df.empty:
        return
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B2"
    _write_df_to_sheet(ws, df, start_row=1, highlight_neg_cols=highlight_cols or [])
    _auto_width(ws)


def _build_news_sheet(wb, df: pd.DataFrame):
    if df is None or df.empty:
        return
    ws = wb.create_sheet("新闻分析")
    ws.sheet_view.showGridLines = False

    # 只保留关键列
    keep = []
    for kw in ["标题", "时间", "来源", "情感标签", "情感分数", "关键词"]:
        col = next((c for c in df.columns if kw in c), None)
        if col:
            keep.append(col)
    export_df = df[keep] if keep else df

    _write_df_to_sheet(ws, export_df, start_row=1,
                       highlight_neg_cols=["情感分数"])
    _auto_width(ws)
    ws.column_dimensions["A"].width = 50  # 标题列加宽


# ── 公开导出函数 ──────────────────────────────────────────────────────────────

def export_to_excel(
    code: str,
    stock_name: str,
    quote_items: list[tuple],   # [("最新价", "10.5"), ...]
    score_items: list[tuple],   # [("综合评分", "75"), ...]
    valuation_items: list[tuple],
    income_df: pd.DataFrame | None = None,
    balance_df: pd.DataFrame | None = None,
    cashflow_df: pd.DataFrame | None = None,
    news_df: pd.DataFrame | None = None,
) -> bytes:
    """
    生成 Excel 报告，返回字节流（供 Streamlit 下载按钮使用）
    """
    wb = openpyxl.Workbook()

    # 摘要页
    _build_summary_sheet(wb, stock_name, code,
                          quote_items, score_items, valuation_items)

    # 财务报表页
    yoy_cols = []  # 同比列高亮
    if income_df is not None:
        yoy_cols_inc = [c for c in income_df.columns if "同比" in c]
        _build_financial_sheet(wb, income_df, "利润表", highlight_cols=yoy_cols_inc)

    if balance_df is not None:
        yoy_cols_bs = [c for c in balance_df.columns if "同比" in c]
        _build_financial_sheet(wb, balance_df, "资产负债表", highlight_cols=yoy_cols_bs)

    if cashflow_df is not None:
        yoy_cols_cf = [c for c in cashflow_df.columns if "同比" in c]
        _build_financial_sheet(wb, cashflow_df, "现金流量表", highlight_cols=yoy_cols_cf)

    # 新闻分析页
    if news_df is not None:
        _build_news_sheet(wb, news_df)

    # 写入内存字节流
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
